import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def load_file(label, book_var):
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if filepath:
        label.config(text=f"Выбранный файл: {filepath}")
        try:
            excel_file = pd.ExcelFile(filepath)
            book_var.set(excel_file.sheet_names[0] if excel_file.sheet_names else '')
            book_combobox['values'] = excel_file.sheet_names
            book_combobox.current(0)
            return filepath
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл: {str(e)}")
    return None

def select_first_file():
    global first_file_path
    first_file_path = load_file(first_file_label, first_book_var)
    update_rcenter_list()

def select_second_file():
    global second_file_path
    second_file_path = load_file(second_file_label, second_book_var)

def update_rcenter_list():
    if not first_file_path:
        update_status("Ошибка: Выберите первую таблицу для загрузки распределительных центров.", "red")
        return
    
    try:
        sheet_name = first_book_var.get()
        df = pd.read_excel(first_file_path, sheet_name=sheet_name)
        if 'Распределительный Центр' in df.columns:
            rcenters = sorted(df['Распределительный Центр'].unique().tolist())
            rcenter_combo['values'] = rcenters
            if rcenters:
                rcenter_combo.current(0)
                update_dates_list()
                update_status("Данные распределительных центров загружены.", "green")
        else:
            update_status("Ошибка: В первой таблице не найдена колонка 'Распределительный Центр'.", "red")
    except Exception as e:
        update_status(f"Ошибка при загрузке данных: {str(e)}", "red")

def update_dates_list():
    rcenter = rcenter_combo.get()
    if not rcenter:
        return

    try:
        sheet_name = first_book_var.get()
        df = pd.read_excel(first_file_path, sheet_name=sheet_name)
        filtered_df = df[df['Распределительный Центр'] == rcenter]

        if 'Дата' in filtered_df.columns:
            dates = sorted(filtered_df['Дата'].dropna().unique().tolist())
            if dates:
                date_combobox['values'] = [date.strftime('%Y-%m-%d') for date in dates]
                if dates:
                    date_combobox.current(0)
                update_status("Даты обновлены.", "green")
        else:
            update_status("Ошибка: В первой таблице не найдена колонка 'Дата'.", "red")
    except Exception as e:
        update_status(f"Ошибка при загрузке дат: {str(e)}", "red")

def process_data():
    try:
        if not first_file_path or not second_file_path:
            update_status("Ошибка: Выберите оба файла.", "red")
            return

        sheet_name_src = first_book_var.get()
        sheet_name_tgt = second_book_var.get()

        df_source = pd.read_excel(first_file_path, sheet_name=sheet_name_src)
        df_target = pd.read_excel(second_file_path, sheet_name=sheet_name_tgt)

        date = date_combobox.get()
        rcenter = rcenter_combo.get()

        df_source['Дата'] = pd.to_datetime(df_source['Дата'], format='%Y-%m-%d', errors='coerce').dt.date
        filtered_data = df_source[(df_source['Распределительный Центр'] == rcenter) & (df_source['Дата'] == pd.to_datetime(date).date())]

        if filtered_data.empty:
            update_status("Ошибка: Нет данных для выбранного распределительного центра и даты.", "red")
            return

        if 'Количество паллет' not in filtered_data.columns:
            update_status("Ошибка: В первой таблице отсутствует колонка 'Количество паллет'.", "red")
            return

        if 'Дата' not in df_target.columns or 'Распределительный Центр' not in df_target.columns:
            update_status("Ошибка: В целевой таблице отсутствуют необходимые колонки.", "red")
            return

        df_target['Дата'] = pd.to_datetime(df_target['Дата'], format='%Y-%m-%d', errors='coerce').dt.date

        workbook = load_workbook(second_file_path)
        sheet = workbook[sheet_name_tgt]

        existing_data_index = df_target[(df_target['Распределительный Центр'] == rcenter) & (df_target['Дата'] == pd.to_datetime(date).date())].index

        fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        if not existing_data_index.empty:
            for index in existing_data_index:
                row_num = index + 2
                pallets_cell = sheet.cell(row=row_num, column=df_target.columns.get_loc('Количество паллет') + 1)

                existing_pallets = df_target.loc[index, 'Количество паллет']
                new_pallets = filtered_data['Количество паллет'].sum()

                if existing_pallets == new_pallets:
                    pallets_cell.fill = fill_green
                else:
                    pallets_cell.fill = fill_red
                    update_status("Ошибка: Количество паллет не совпадает!", "red")
        else:
            new_rows = []
            for _, row in filtered_data.iterrows():
                new_rows.append([row['Дата'], row['Распределительный Центр'], row['Количество паллет']])

            for new_row in new_rows:
                sheet.append(new_row)

            update_status("Запрос обработан", "green")

        df_target = pd.concat([df_target, filtered_data], ignore_index=True)
        df_target = df_target.drop_duplicates(subset=['Распределительный Центр', 'Дата'], keep='last')
        df_target.to_excel(second_file_path, sheet_name=sheet_name_tgt, index=False)

        workbook.save(second_file_path)

    except Exception as e:
        update_status(f"Произошла ошибка: {str(e)}", "red")

def update_status(message, color):
    status_label.config(text=message, fg=color)
    root.update_idletasks()  # Обновляем интерфейс

# Создаем GUI
root = tk.Tk()
root.title("Перенос данных между таблицами")

# Поля для выбора файлов
first_file_label = tk.Label(root, text="Файл с исходными данными не выбран")
first_file_label.pack(pady=5)
first_file_button = tk.Button(root, text="Выбрать первую таблицу", command=select_first_file)
first_file_button.pack(pady=5)

second_file_label = tk.Label(root, text="Файл для вставки данных не выбран")
second_file_label.pack(pady=5)
second_file_button = tk.Button(root, text="Выбрать вторую таблицу", command=select_second_file)
second_file_button.pack(pady=5)

# Поле для выбора книги (таблицы) в первом и втором файле
tk.Label(root, text="Выберите книгу из первого файла:").pack(pady=5)
first_book_var = tk.StringVar()
book_combobox = ttk.Combobox(root, textvariable=first_book_var)
book_combobox.pack(pady=5)

tk.Label(root, text="Выберите книгу из второго файла:").pack(pady=5)
second_book_var = tk.StringVar()
book_combobox2 = ttk.Combobox(root, textvariable=second_book_var)
book_combobox2.pack(pady=5)

# Поле для выбора распределительного центра
tk.Label(root, text="Выберите распределительный центр:").pack(pady=5)
rcenter_combo = ttk.Combobox(root)
rcenter_combo.pack(pady=5)
rcenter_combo.bind("<<ComboboxSelected>>", lambda e: update_dates_list())

# Поле для выбора даты
tk.Label(root, text="Выберите дату:").pack(pady=5)
date_combobox = ttk.Combobox(root)
date_combobox.pack(pady=5)

# Кнопка для обработки данных
process_button = tk.Button(root, text="Обработать данные", command=process_data)
process_button.pack(pady=20)

# Статус обработки
status_label = tk.Label(root, text="", font=("Arial", 12))
status_label.pack(pady=10)

root.mainloop()
